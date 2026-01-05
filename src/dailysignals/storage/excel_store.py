from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook

from src.dailysignals.constants import SHEET_READINGS, READING_COLUMNS


def ensure_workbook(excel_path: Path) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        wb = load_workbook(excel_path)
        if SHEET_READINGS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_READINGS)
            ws.append(READING_COLUMNS)
            wb.save(excel_path)
            return

        ws = wb[SHEET_READINGS]
        if ws.max_row < 1:
            ws.append(READING_COLUMNS)
            wb.save(excel_path)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_READINGS
    ws.append(READING_COLUMNS)
    wb.save(excel_path)


def _build_index(ws) -> Dict[Tuple[str, str], int]:
    """
    Build an index: (date, signal_id) -> row_number (1-based in Excel)
    Assumes header is row 1.
    """
    col_date = READING_COLUMNS.index("date") + 1
    col_signal = READING_COLUMNS.index("signal_id") + 1

    idx: Dict[Tuple[str, str], int] = {}
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=col_date).value
        sid = ws.cell(row=r, column=col_signal).value
        if d and sid:
            idx[(str(d), str(sid))] = r
    return idx


def upsert_readings(excel_path: Path, rows: List[Dict[str, Any]]) -> Dict[str, int]:
    """
    Enforces one reading per (date, signal_id).
    If exists -> replace row values; else -> append new row.

    Returns counts: {"inserted": x, "updated": y}
    """
    ensure_workbook(excel_path)
    wb = load_workbook(excel_path)
    ws = wb[SHEET_READINGS]

    index = _build_index(ws)
    now = datetime.now().isoformat(timespec="seconds")

    inserted = 0
    updated = 0

    for r in rows:
        r = dict(r)
        r.setdefault("created_at", now)

        key = (str(r.get("date", "")), str(r.get("signal_id", "")))
        if not key[0] or not key[1]:
            continue  # ignore malformed rows

        existing_row = index.get(key)
        if existing_row:
            # Update in place
            for c, col_name in enumerate(READING_COLUMNS, start=1):
                ws.cell(row=existing_row, column=c).value = r.get(col_name, "")
            updated += 1
        else:
            ws.append([r.get(col, "") for col in READING_COLUMNS])
            inserted += 1
            index[key] = ws.max_row

    wb.save(excel_path)
    return {"inserted": inserted, "updated": updated}
